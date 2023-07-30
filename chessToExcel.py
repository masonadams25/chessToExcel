from chessdotcom import get_player_games_by_month
import openpyxl
from time import strftime, localtime

username = "Masonrules9"
month = "07"
year = "2023"
response = get_player_games_by_month(username = username, year = year, month = month)
games = response.json['games']

dailyGames = {}
def addDailyGames(dict, result, month, day, rating, accuracy):
    dict['date'] = str(month) + '-' + str(day)
    dict['totGames'] = int(dict['totGames']) + 1
    dict['avgRating'] = (int(dict['avgRating']) * (int(dict['totGames'])-1) + int(rating)) / int(dict['totGames'])
    if result == 'win':
        dict['wins'] = int(dict['wins'] + 1)
    else:
        dict['losses'] = int(dict['losses'] + 1)
    dict['winPercent'] =  (float(dict['wins']) / float(dict['totGames']))*100
    if accuracy != 'N/A':
        dict['avgAccuracy'] = (int(dict['avgAccuracy']) * (int(dict['totGames'])-1) + int(accuracy)) / int(dict['totGames'])

def writeToExcel(dict):
    path = "chessToExcel.xlsx"
    try:
        wb = openpyxl.load_workbook(path)
        print("Excel file loaded")
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        print("Excel file created")
        print("Path: " + path)
    sheet = wb.active

    sheet.cell(row = 1, column = 1).value = "Date"
    sheet.cell(row = 2, column = 1).value = "Avg Rating"
    sheet.cell(row = 3, column = 1).value = "Games"
    sheet.cell(row = 4, column = 1).value = "Wins"
    sheet.cell(row = 5, column = 1).value = "Losses"
    sheet.cell(row = 6, column = 1).value = "Win %"
    sheet.cell(row = 7, column = 1).value = "Avg acc"

    print("Excel file saved")

    count = 2
    for day in dict:
        sheet.cell(row = 1, column = count).value = str(dict[day]['date'])
        sheet.cell(row = 2, column = count).value = str(round(float(dailyGames[day]['avgRating']),2))
        sheet.cell(row = 3, column = count).value = str(dailyGames[day]['totGames'])
        sheet.cell(row = 4, column = count).value = str(dailyGames[day]['wins'])
        sheet.cell(row = 5, column = count).value = str(dailyGames[day]['losses'])
        sheet.cell(row = 6, column = count).value = str(round(float(dailyGames[day]['winPercent']),2))
        sheet.cell(row = 7, column = count).value = str(round(float(dailyGames[day]['avgAccuracy']),2))
        
        count = count + 1

    wb.save(path)
        

    
for game in games:
    if game['white']['username'] == username:
        color = 'white'
    else:
        color = 'black'   

    result = game[color]['result']
    date = strftime('%d-%m', localtime(game['end_time']))
    month = date[-2:]
    day = date[:2]
    rating = game[color]['rating']
    try:
        accuracy = game['accuracies'][color]
    except KeyError:
        accuracy = 'N/A'


    try:
        dailyGames[date]
    except KeyError:
        dailyGames[date] = {
            'date': 0,
            'avgRating': 0,
            'totGames': 0,
            'wins': 0,
            'losses': 0,
            'winPercent': 0,
            'avgAccuracy': 0
        }
    addDailyGames(dailyGames[date], result, month, day, rating, accuracy)

    #print(str(result).ljust(10) + " " + str(accuracy).ljust(6) + " " + str(rating).ljust(4) + " " + str(date).ljust(10))

writeToExcel(dailyGames)
for day in dailyGames:
    print(str(dailyGames[day]['date']).ljust(6) + 
          "- Avg rating: " + str(round(float(dailyGames[day]['avgRating']),2)).ljust(7) +
          "- Games: " + str(dailyGames[day]['totGames']).ljust(3) +
          "- Wins: " + str(dailyGames[day]['wins']).ljust(3) +
          "- Losses: " + str(dailyGames[day]['losses']).ljust(3) +
          "- Win %: " + str(round(float(dailyGames[day]['winPercent']),2)).ljust(6) +
          "- Avg acc: " + str(round(float(dailyGames[day]['avgAccuracy']),2)).ljust(7))


